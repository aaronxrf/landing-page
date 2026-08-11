#!/usr/bin/env python3
"""Build docs/doctors.json from all *_gimenes_arsti*.xlsx files.

Region is taken from the sheet name (not the filename), header is row 8 and
data rows are 9..max_row-2. Files whose sheet region was already processed are
skipped, so duplicate publications (e.g. NVD's Vidzeme file, which is a copy
of Kurzeme) collapse.

Geocodes unique practice addresses via Nominatim (keyless), cached in
geocode_cache.json so rebuilds skip the network.
"""
import glob
import json
import os
import re
import time
import urllib.parse
import urllib.request

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
CACHE = os.path.join(HERE, "geocode_cache.json")
OUT = os.path.join(HERE, "docs", "doctors.json")
USER_AGENT = "gimenes-arsti-map/0.1 (localhost; personal use)"

DAYS = ["Pr", "O", "T", "C", "Pk", "S", "Sv"]


def region_files():
    seen = set()
    for path in sorted(glob.glob(os.path.join(HERE, "*_gimenes_arsti_*.xlsx"))):
        wb = openpyxl.load_workbook(path, read_only=True)
        region = wb.sheetnames[0]
        wb.close()
        if region in seen:
            print(f"skip {path}: duplicate region {region}")
            continue
        seen.add(region)
        yield path, region


def load_doctors(path, region):
    wb = openpyxl.load_workbook(path, data_only=True)
    wbf = openpyxl.load_workbook(path, data_only=False)
    ws, wsf = wb[region], wbf[region]
    start = 8  # legacy format: no header row, data from row 8
    for r in range(1, 13):
        if ws.cell(r, 4).value == "Vārds, uzvārds":
            start = r + 1  # new format: header row, data follows
            break
    doctors = []
    for r in range(start, ws.max_row + 1):
        name = ws.cell(r, 4).value
        if not name:
            continue
        open_flag = wsf.cell(r, 1).fill.patternType is not None
        hours_reception = " · ".join(
            f"{DAYS[i]} {ws.cell(r, 23 + i).value}"
            for i in range(7)
            if ws.cell(r, 23 + i).value
        )
        hours_practice = " · ".join(
            f"{DAYS[i]} {ws.cell(r, 16 + i).value}"
            for i in range(7)
            if ws.cell(r, 16 + i).value
        )
        doctors.append({
            "name": str(name).strip(),
            "practice": str(ws.cell(r, 3).value or "").strip(),
            "specialty": str(ws.cell(r, 7).value or "").strip(),
            "address": str(ws.cell(r, 10).value or "").strip(),
            "phone": str(ws.cell(r, 11).value or "").strip(),
            "email": str(ws.cell(r, 12).value or "").strip(),
            "vaccination": str(ws.cell(r, 5).value or "").strip(),
            "patients": str(ws.cell(r, 15).value or "").strip(),
            "hours_reception": hours_reception,
            "hours_practice": hours_practice,
            "open": open_flag,
            "region": region,
        })
    wb.close()
    wbf.close()
    return doctors


def nominatim(q):
    url = "https://nominatim.openstreetmap.org/search?" + urllib.parse.urlencode({
        "q": q, "format": "jsonv2", "limit": 3, "accept-language": "lv",
    })
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.load(resp)
    for hit in data:
        if hit.get("class") not in ("highway", "boundary", "waterway", "railway"):
            return float(hit["lat"]), float(hit["lon"])
    return None


def candidate_queries(address):
    base = re.sub(r",\s*LV[- ]?\d{4}$", "", address).strip().rstrip(",")
    yield base + ", Latvija"
    parts = [p.strip() for p in base.split(",") if p.strip()]
    street = re.match(r".+?\d+[A-Za-z]?", parts[0])
    street = (street.group() if street else parts[0]).strip().strip('"')
    town = next((p for p in parts[1:]
                 if not any(k in p for k in ("novads", "pagasts", "priekšpilsēta"))), None)
    city = "Rīga" if any("Rīga" in p for p in parts) else town
    if city:
        yield f"{street}, {city}, Latvija"
        yield f"{street}, Latvija"
    if town and town != city:
        yield f"{town}, Latvija"


def geocode(address, cache):
    if address in cache:
        return cache[address]
    for q in candidate_queries(address):
        hit = nominatim(q)
        if hit:
            cache[address] = hit
            save_cache(cache)
            return hit
        time.sleep(1.1)
    return None


def load_cache():
    try:
        with open(CACHE) as f:
            return json.load(f)
    except (OSError, json.JSONDecodeError):
        return {}


def save_cache(cache):
    with open(CACHE, "w") as f:
        json.dump(cache, f, ensure_ascii=False, indent=1)


def main():
    doctors = []
    for path, region in region_files():
        docs = load_doctors(path, region)
        print(f"{path}: {len(docs)} doctors ({region})")
        doctors.extend(docs)
    print(f"{len(doctors)} doctors total")

    by_addr = {}
    for d in doctors:
        by_addr.setdefault(d["address"], []).append(d)
    print(f"{len(by_addr)} unique addresses")

    cache = load_cache()
    locations, failed = [], []
    for i, (address, docs) in enumerate(sorted(by_addr.items()), 1):
        if address not in cache:
            print(f"[{i}/{len(by_addr)}] {address}")
            time.sleep(1.1)
        pos = geocode(address, cache)
        if pos is None:
            failed.append(address)
            continue
        locations.append({
            "lat": pos[0], "lng": pos[1], "address": address,
            "doctors": sorted(docs, key=lambda d: d["name"]),
        })
    locations.sort(key=lambda loc: (loc["lat"], loc["lng"]))

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w") as f:
        json.dump({"locations": locations}, f, ensure_ascii=False)
    print(f"wrote {OUT}: {len(locations)} locations")
    if failed:
        print("FAILED geocode:", *failed, sep="\n  ")


if __name__ == "__main__":
    main()
